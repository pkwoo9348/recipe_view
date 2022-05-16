import openpyxl

dir = r"C:\Users\kp\Desktop\recipe_view\test.xlsx"

# 경로에 있는 excel파일을 불러온다.
excel = openpyxl.load_workbook(dir)

# create_sheet() 메소드로 새로운 시트를 생성한다.
excel.create_sheet('chocolate')
excel.create_sheet('bread')

excel.save(dir)
excel = openpyxl.load_workbook(dir)

# bread 시트를 선택
excel_ws = excel['bread']

# A1 셀을 직접 지정해 입력
excel_ws['A1'] = 'HONEY'

# 셀을 행과 열로 지정해 입력
excel_ws.cell(row = 2, column = 1).value = 'Cheese'

# 입력되어있는 바로 밑 행의 처음부터 입력된다.
for x in range(10):
    excel_ws.append(['I', 'like', 'bread'])

excel.save(dir)