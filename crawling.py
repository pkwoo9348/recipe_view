# 셀레니움으로 모든 동영상 url 가져오기
# 셀레니움으로 url순회하며 크롤링, 리스트에 append
# 엑셀에 저장

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.by import By
import openpyxl as xl 


# selenium에서 사용할 웹 드라이버 절대 경로 정보
chromedriver = r'C:\Users\kp\Desktop\recipe_view\chromedriver_win32\chromedriver.exe'
# selenum의 webdriver에 앞서 설치한 chromedirver를 연동한다.
driver = webdriver.Chrome(chromedriver)
# driver로 특정 페이지를 크롤링한다.
driver.get('https://www.youtube.com/c/paikscuisine/videos')
driver.implicitly_wait(5)

#스크롤 마지막까지 내리기
before_h = driver.execute_script("return window.scrollY")
while True:
    driver.find_element_by_css_selector('body').send_keys(Keys.END)
    time.sleep(2)
    
    after_h=driver.execute_script("return window.scrollY")
    
    if after_h==before_h:
        break
    
    before_h = after_h

excel_url=[]    # 엑셀에 저장할 url
excel_coment=[] # 엑셀에 저장할 coment

#url 가져오기
url_lists=driver.find_elements_by_css_selector("a.yt-simple-endpoint.style-scope.ytd-grid-video-renderer")

# url 순회 데이터 리스트에 append
for url_list in url_lists:
    url=url_list.get_attribute("href")
    driver2= webdriver.Chrome(chromedriver)
    driver2.get(url)
    driver2.implicitly_wait(5)
    coments=driver2.find_elements(By.CSS_SELECTOR, 'span.style-scope.yt-formatted-string')
    driver2.implicitly_wait(5)

    # coment 문자열 길이가 100이 넘는 문자열만 url과 함께 저장
    for coment in coments:
        if len(coment.text)> 100:
            excel_url.append(url)
            excel_coment.append(coment.text)
            break
        else:
            driver2.close
            continue


# xlsx 파일 생성
wb = xl.Workbook()
sheet = wb.active
sheet.title = '테스트'
# 컬럼명 지정(헤더)
col_names = ['url', 'coment'] 

for seq, name in enumerate(col_names): 
    sheet.cell(row=1, column=seq+1, value=name) 

wb.save(r"C:\Users\kp\Desktop\recipe_view\test.xlsx") 
wb.close()

time.sleep(50)

# 엑셀에 저장
dir = r"C:\Users\kp\Desktop\recipe_view\test.xlsx"

excel = xl.load_workbook(dir)

# 시트 선택
excel_ws = excel['테스트']
i=0
j=len(excel_url)
# 입력되어있는 바로 밑 행의 처음부터 입력된다.
while i<j:
    excel_ws[f'A{i+2}'] = excel_url[i]
    excel_ws[f'B{i+2}'] = excel_coment[i]
    i=i+1


excel.save(dir)
